import sys
sys.path.append(".")
import os
from PyQt6 import QtWidgets, QtGui, uic
from PyQt6.QtCore import Qt, QEvent
from PyQt6.QtWidgets import QFileDialog
import openpyxl, re
        
class TrackModelModule(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        uic.loadUi('Modules/Track_Model/Frontend/Track_Model_UI.ui',self)
        
        # disable line selector until track is loaded to avoid undefined behavior
        self.TrackLineColorValue.setEnabled(False)
        
        # line color combo box hint behavior, hide hint initially
        self.TrackLineColorValue.installEventFilter(self)
        self.LineSelectHint.setVisible(True)
        
        # load track model button
        self.LoadTrackModelButton.clicked.connect(self.track_layout)

        # set up scene for graphics view and set scene size to widget size 
        self.graphicsView.setScene(QtWidgets.QGraphicsScene())
        self.graphicsView.setSceneRect(0,0,1221,521)
        
        # build track in graphics view
        self.TrackLineColorValue.currentTextChanged.connect(self.build_track_map)
        
    # functions 
    
    def track_layout(self):
        file_filter = 'Excel File (*.xlsx)'
        response, _ = QFileDialog.getOpenFileName(
            parent=self,
            caption = 'Select a Track Layout',
            directory=os.getcwd(),
            filter=file_filter,
            initialFilter='Execl File (*.xlsx)',
        )
        
        # parse layout file
        file_name = os.path.basename(response) # extract the filename from the path
        layout_data = openpyxl.load_workbook(file_name, data_only=True) # open workbook and extract data only (take result from cells with formulas)
        
        # get assign sheets to sheet names from the file
        sheet3 = layout_data['Red Line']
        sheet4 = layout_data['Green Line']
        
        # declare list to store line data
        self.red_line_data = []
        self.green_line_data = []
        
        # iterate through to extract data
        for row in sheet3.iter_rows(min_row=2, max_row=77, values_only=True):
            line, section, block_number, block_length, block_grade, speed_limit, infrastructure, station_side, elevation, cumulative_elevation = row[:10]
            
            self.red_line_data.append((line, section, block_number, block_length, block_grade, speed_limit, infrastructure, station_side, elevation, cumulative_elevation))
            
        
        for row in sheet4.iter_rows(min_row=2,max_row=151, values_only=True):
            line, section, block_number, block_length, block_grade, speed_limit, infrastructure, station_side, elevation, cumulative_elevation, traversal_time = row[:11]
            
            self.green_line_data.append((line, section, block_number, block_length, block_grade, speed_limit, infrastructure, station_side, elevation, cumulative_elevation, traversal_time))
        
        # enable line selection after track layout is loaded 
        self.LineSelectHint.setVisible(False)
        self.TrackLineColorValue.setEnabled(True)
        
        
        #TODO return lists from this function 
    
    def show_block_info(self,block_number):
        self.block_length_display.setText(str(self.red_line_data[block_number][3]))
            
        
    def build_track_map(self):
        line_name = self.TrackLineColorValue.currentText()
    
        if line_name == 'Red Line':
            # place the yard block and go from there
            block0 = QtWidgets.QGraphicsRectItem(800,0,40,40)
            block0.setBrush(QtGui.QColor(128,128,128)) # gray color for yard block 
            self.graphicsView.scene().addItem(block0)
            
            # add label to yard 
            text = QtWidgets.QGraphicsTextItem('Yard')
            text.setPos(802,40)
            self.graphicsView.scene().addItem(text)
            
            # use function to keep building map
            self.add_block_to_map(800,-70,40,'block9','9','right',800,0)
            self.add_block_to_map(775,-120,40,'block8','8','right',800,-70)
            self.add_block_to_map(740,-170,40,'block7','7','top',775,-120)
            self.add_block_to_map(690,-170,40,'block6','6','top',740,-170)
            self.add_block_to_map(640,-150,40,'block5','5','top',690,-170)
            self.add_block_to_map(590,-130,40,'block4','4','top',640,-150)
            self.add_block_to_map(540,-110,40,'block3','3','top',590,-130)
            self.add_block_to_map(490,-90,40,'block2','2','top',540,-110)
            self.add_block_to_map(440,-70,40,'block1','1','top',490,-90)
            self.add_block_to_map(390,-30,40,'block16','16','top',440,-70)
            self.add_block_to_map(460,-10,40,'block15','15','bottom',390,-30)
            self.add_block_to_map(520,-12,40,'block14','14','bottom',460,-10)
            self.add_block_to_map(580,-14,40,'block13','13','bottom',520,-12)
            self.add_block_to_map(630,-16,40,'block12','12','bottom',580,-14)
            self.add_block_to_map(680,-20,40,'block11','11','bottom',630,-16)
            self.add_block_to_map(730,-24,40,'block10','10','bottom',680,-20)
            
            # connect blocks 9 and 10
            line = QtWidgets.QGraphicsLineItem(770,-4,800,-50)
            line.setPen(QtGui.QColor(255,255,255))
            self.graphicsView.scene().addItem(line)
            
            # add more blocks
            self.add_block_to_map(330,-30,40,'block17','17','top',390,-30)
            self.add_block_to_map(270,-30,40,'block18','18','top',330,-30)
            self.add_block_to_map(210,-30,40,'block19','19','top',270,-30)
            self.add_block_to_map(150,-30,40,'block20','20','top',210,-30)
            self.add_block_to_map(90,-20,40,'block21','21','top',150,-30)
            self.add_block_to_map(40,30,40,'block22','22','left',90,-20)
            self.add_block_to_map(30,80,40,'block23','23','left',40,30)
            self.add_block_to_map(60,130,40,'block24','24','left',30,80)
            self.add_block_to_map(105,140,40,'block25','25','top',60,130)
            self.add_block_to_map(150,150,40,'block26','26','top',105,140)
            self.add_block_to_map(195,160,40,'block27','27','top',150,150)
            self.add_block_to_map(240,170,40,'block28','28','top',195,160)
            self.add_block_to_map(285,180,40,'block29','29','top',240,170)
            self.add_block_to_map(330,190,40,'block30','30','top',285,180)
            self.add_block_to_map(375,200,40,'block31','31','top',330,190)
            self.add_block_to_map(420,210,40,'block32','32','top',375,200)
            self.add_block_to_map(465,220,40,'block33','33','top',420,210)
            self.add_block_to_map(510,220,40,'block34','34','top',465,220)
            self.add_block_to_map(555,220,40,'block35','35','top',510,220)
            self.add_block_to_map(600,220,40,'block36','36','top',555,220)
            self.add_block_to_map(645,220,40,'block37','37','top',600,220)
            self.add_block_to_map(690,220,40,'block38','38','top',645,220)
            self.add_block_to_map(735,220,40,'block39','39','top',690,220)
            self.add_block_to_map(780,220,40,'block40','40','top',735,220)
            self.add_block_to_map(825,220,40,'block41','41','top',780,220)
            self.add_block_to_map(870,265,40,'block42','42','right',825,220)
            self.add_block_to_map(915,310,40,'block43','43','right',870,265)
            
        elif line_name == 'Green Line':
            # place the yard block 
            block0 = QtWidgets.QGraphicsRectItem(900,180,20,20)
            block0.setBrush(QtGui.QColor(128,128,128)) # gray color for yard block 
            self.graphicsView.scene().addItem(block0)
            
            # add label to yard 
            text = QtWidgets.QGraphicsTextItem('Yard')
            text.setPos(892,160)
            self.graphicsView.scene().addItem(text)
            
            # continue building map using function 
            self.add_block_to_map(900,250,20,'block63','63','right',900,180)
            self.add_block_to_map(900,290,20,'block64','64','right',900,250)
            self.add_block_to_map(900,330,20,'block65','65','right',900,290)
            self.add_block_to_map(900,370,20,'block66','66','right',900,330)
            self.add_block_to_map(900,410,20,'block67','67','right',900,370)
            self.add_block_to_map(900,450,20,'block68','68','right',900,410)
            self.add_block_to_map(900,490,20,'block69','69','right',900,450)
            self.add_block_to_map(900,530,20,'block70','70','right',900,490)
            self.add_block_to_map(900,570,20,'block71','71','right',900,530)
            self.add_block_to_map(900,610,20,'block72','72','right',900,570)
            self.add_block_to_map(850,640,20,'block73','73','right',880,610)
            self.add_block_to_map(810,640,20,'block74','74','bottom',850,640)
            self.add_block_to_map(770,640,20,'block75','75','bottom',810,640)
            self.add_block_to_map(730,640,20,'block76','76','bottom',770,640)
            self.add_block_to_map(690,640,20,'block77','77','bottom',730,640)
            self.add_block_to_map(650,640,20,'block78','78','bottom',690,640)
            self.add_block_to_map(610,640,20,'block79','79','bottom',650,640)
            self.add_block_to_map(570,640,20,'block80','80','bottom',610,640)
            self.add_block_to_map(530,640,20,'block81','81','bottom',570,640)
            self.add_block_to_map(490,640,20,'block82','82','bottom',530,640)
            self.add_block_to_map(450,640,20,'block83','83','bottom',490,640)
            self.add_block_to_map(410,640,20,'block84','84','bottom',450,640)
            self.add_block_to_map(370,640,20,'block85','85','bottom',410,640)
            self.add_block_to_map(330,640,20,'block86','86','bottom',370,640)
            self.add_block_to_map(290,640,20,'block87','87','bottom',330,640)
        
        block0.mousePressEvent = self.show_block_info(0)

    def add_block_to_map(self,x,y,block_size,block_number,block_number_2,label_pos,prev_x,prev_y):
            block_number = QtWidgets.QGraphicsRectItem(x,y,block_size,block_size)
            block_number.setBrush(QtGui.QColor(0,128,0)) # gray color for yard block 
            self.graphicsView.scene().addItem(block_number)
            
            # handle different positions for the block label
            if label_pos == 'top' and y<0:
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x+10,y-(block_size/2))
                self.graphicsView.scene().addItem(text)
            
            if label_pos == 'top' and y>=0:
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x+10,y-(block_size/2))
                self.graphicsView.scene().addItem(text)
            
            if label_pos == 'right':
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x+block_size,y)
                self.graphicsView.scene().addItem(text)
            
            if label_pos == 'left':
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x-(block_size/2),y+10)
                self.graphicsView.scene().addItem(text)
            
            if label_pos == 'bottom' and y<0:
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x+(block_size*(2/3)),y+block_size)
                self.graphicsView.scene().addItem(text)
            
            if label_pos == 'bottom' and y>=0:
                text = QtWidgets.QGraphicsTextItem(block_number_2)
                text.setPos(x,y+block_size)
                self.graphicsView.scene().addItem(text)
            
            # draw line to connect blocks with handling for different orientations 
            # vertically aligned (negative)
            if (abs(prev_x-x)) <= 10 and y<0: 
                line = QtWidgets.QGraphicsLineItem(x+(block_size/2),prev_y,prev_x+(block_size/2),y+block_size)
                line.setPen(QtGui.QColor(255,255,255)) # white color for lines
                self.graphicsView.scene().addItem(line)
            
            # vertically aligned (positive)
            if (abs(prev_x-x)) <= 30 and y>=0: 
                line = QtWidgets.QGraphicsLineItem((x+(block_size/2)),y,prev_x+(block_size/2),prev_y+block_size)
                line.setPen(QtGui.QColor(255,255,255)) # white color for lines
                self.graphicsView.scene().addItem(line)
                return 
            
            if x == prev_x and y>=0:
                line = QtWidgets.QGraphicsLineItem(x+(block_size*(8/3)),y,prev_x+(block_size/2),prev_y+block_size)
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            
            # blocks in line horizontally
            if y == prev_y and y<0:
                line = QtWidgets.QGraphicsLineItem(x+block_size,prev_y+(block_size/2),prev_x,prev_y+(block_size/2))
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            
            if y == prev_y and y>=0:
                line = QtWidgets.QGraphicsLineItem(x+block_size,prev_y+(block_size/2),prev_x,prev_y+(block_size/2))
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
                return
             
            if (abs(prev_y-y)) <= 20 and y>=0:
                line = QtWidgets.QGraphicsLineItem(x,y+20,prev_x+40,prev_y+25)
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
                return
            
            # diagonal cases
            # up to the left
            elif x != prev_x and y != prev_y and x<prev_x and y<prev_y and y<0:
                line = QtWidgets.QGraphicsLineItem(x+(block_size-10),y+block_size,prev_x+(block_size/4),prev_y)
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            
            # down to the left
            elif x != prev_x and y != prev_y and x<prev_x and y>prev_y:
                line = QtWidgets.QGraphicsLineItem(prev_x,prev_y+(block_size/2),x+block_size,y+(block_size/2))
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            
            # down to the right
            elif x != prev_x and y != prev_y and prev_x<x and y>prev_y:
                line = QtWidgets.QGraphicsLineItem(prev_x+block_size,prev_y+(block_size/2),x,y+(block_size/2))
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            
            # up to the right (negative y-axis)
            elif x != prev_x and y != prev_y and x>prev_x and y<prev_y and y<0:
                line = QtWidgets.QGraphicsLineItem(prev_x+block_size,prev_y+(block_size/2),x,y+(block_size/2))
                line.setPen(QtGui.QColor(255,255,255))
                self.graphicsView.scene().addItem(line)
            

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = TrackModelModule()
    window.show()
    sys.exit(app.exec())
